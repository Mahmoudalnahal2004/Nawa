import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Questions Template"

# Define headers
headers = [
    "Question Text",
    "Option A",
    "Option B",
    "Option C",
    "Option D",
    "Option E",
    "Correct Answer",
    "Explanation",
    "Difficulty",
    "Category Name"
]

# Write headers and apply styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

for col_num, header_title in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header_title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    if col_num in [1, 8]:  # Question Text and Explanation
        ws.column_dimensions[cell.column_letter].width = 40
    else:
        ws.column_dimensions[cell.column_letter].width = 20

# Add sample row
sample_row = [
    "What is the powerhouse of the cell?",
    "Mitochondria",
    "Nucleus",
    "Ribosome",
    "Endoplasmic Reticulum",
    "",
    "A",
    "Mitochondria generate most of the chemical energy needed to power the cell's biochemical reactions.",
    "Easy",
    "Biology"
]
for col_num, value in enumerate(sample_row, 1):
    ws.cell(row=2, column=col_num, value=value)

# Save the workbook to apps/web/public
output_dir = os.path.join("..", "web", "public")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "question_template.xlsx")

wb.save(output_path)
print(f"Template successfully saved to {os.path.abspath(output_path)}")
