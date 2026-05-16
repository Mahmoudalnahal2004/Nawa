import io
from typing import List

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.question import Question, Difficulty, QuestionStatus
from app.models.category import Category
from app.schemas.question import ImportResult, ImportError_


# Expected column mapping (case-insensitive)
COLUMN_MAP = {
    "question_text": ["question_text", "question", "stem", "q"],
    "option_a": ["option_a", "a", "choice_a"],
    "option_b": ["option_b", "b", "choice_b"],
    "option_c": ["option_c", "c", "choice_c"],
    "option_d": ["option_d", "d", "choice_d"],
    "option_e": ["option_e", "e", "choice_e"],
    "correct_answer": ["correct_answer", "answer", "correct"],
    "explanation": ["explanation", "rationale", "explain"],
    "difficulty": ["difficulty", "level"],
    "category_name": ["category_name", "category", "subject", "module"],
}


def _find_column_index(headers: List[str], field: str) -> int | None:
    """Find the column index for a field using flexible name matching."""
    aliases = COLUMN_MAP.get(field, [field])
    for i, header in enumerate(headers):
        if header and header.strip().lower() in aliases:
            return i
    return None


async def _get_or_create_category(db: AsyncSession, name: str) -> int:
    """Get category ID by name, or create it if it doesn't exist."""
    result = await db.execute(select(Category).where(Category.name == name))
    category = result.scalar_one_or_none()
    if category:
        return category.id

    category = Category(name=name)
    db.add(category)
    await db.flush()
    await db.refresh(category)
    return category.id


async def import_excel(db: AsyncSession, file_content: bytes, default_category_id: int | None = None) -> ImportResult:
    """Parse and import questions from an Excel file.

    Validation rules:
    - Skip rows where question_text or correct_answer is missing
    - Flag rows with invalid correct_answer (must be A-E)
    - All imported questions start as Draft
    """
    wb = load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return ImportResult(imported=0, skipped=0, errors=[
            ImportError_(row=1, reason="File is empty or has no data rows")
        ])

    # Parse headers
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    # Map columns
    col_indices = {}
    for field in COLUMN_MAP:
        idx = _find_column_index(headers, field)
        col_indices[field] = idx

    # Validate required columns exist
    if col_indices["question_text"] is None:
        return ImportResult(imported=0, skipped=0, errors=[
            ImportError_(row=1, reason="Missing required column: question_text")
        ])
    if col_indices["correct_answer"] is None:
        return ImportResult(imported=0, skipped=0, errors=[
            ImportError_(row=1, reason="Missing required column: correct_answer")
        ])

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            # Extract values
            def get_val(field: str) -> str | None:
                idx = col_indices.get(field)
                if idx is None or idx >= len(row):
                    return None
                val = row[idx]
                return str(val).strip() if val is not None else None

            question_text = get_val("question_text")
            correct_answer = get_val("correct_answer")

            # Validation: required fields
            if not question_text:
                errors.append(ImportError_(row=row_idx, reason="Missing question_text"))
                skipped += 1
                continue

            if not correct_answer:
                errors.append(ImportError_(row=row_idx, reason="Missing correct_answer"))
                skipped += 1
                continue

            # Validate correct_answer
            correct_answer = correct_answer.upper().strip()
            if correct_answer not in ["A", "B", "C", "D", "E"]:
                errors.append(ImportError_(row=row_idx, reason=f"Invalid correct_answer: '{correct_answer}' (must be A-E)"))
                skipped += 1
                continue

            # Get options
            option_a = get_val("option_a") or ""
            option_b = get_val("option_b") or ""
            option_c = get_val("option_c") or ""
            option_d = get_val("option_d") or ""
            option_e = get_val("option_e") or ""

            if not option_a or not option_b:
                errors.append(ImportError_(row=row_idx, reason="Must have at least options A and B"))
                skipped += 1
                continue

            # Get optional fields
            explanation = get_val("explanation") or ""
            difficulty_str = (get_val("difficulty") or "medium").lower()
            if difficulty_str not in ["easy", "medium", "hard"]:
                difficulty_str = "medium"

            # Handle category
            category_name = get_val("category_name")
            if category_name:
                category_id = await _get_or_create_category(db, category_name)
            elif default_category_id:
                category_id = default_category_id
            else:
                # Create a default "Uncategorized" category
                category_id = await _get_or_create_category(db, "Uncategorized")

            # Create question
            question = Question(
                category_id=category_id,
                question_text=question_text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                option_e=option_e,
                correct_answer=correct_answer,
                explanation=explanation,
                difficulty=Difficulty(difficulty_str),
                status=QuestionStatus.DRAFT,  # Always start as draft
            )
            db.add(question)
            imported += 1

        except Exception as e:
            errors.append(ImportError_(row=row_idx, reason=f"Unexpected error: {str(e)}"))
            skipped += 1

    if imported > 0:
        await db.flush()

    wb.close()
    return ImportResult(imported=imported, skipped=skipped, errors=errors)
